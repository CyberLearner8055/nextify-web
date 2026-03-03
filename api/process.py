from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
import openpyxl
import requests
import tempfile
from datetime import datetime

app = FastAPI()

API_URL = "https://www.zephrexdigital.site/api?key=ZEPH-IM45D&type=PHONE&term="

def lookup_number(number):
    try:
        r = requests.get(API_URL + str(number), timeout=10)
        data = r.json()
        if data.get("status") and data.get("data"):
            return data["data"][0]
    except:
        return None
    return None

@app.post("/process")
async def process_excel(file: UploadFile = File(...)):
    try:
        temp_in = tempfile.mktemp(".xlsx")
        temp_out = tempfile.mktemp(".xlsx")

        with open(temp_in, "wb") as f:
            f.write(await file.read())

        wb = openpyxl.load_workbook(temp_in)
        ws = wb.active

        headers = ["Number", "Name", "Father Name", "Address", "Email", "Alt Mobile", "Status"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = h

        for row in range(2, ws.max_row + 1):
            num = ws.cell(row=row, column=1).value
            if not num:
                continue

            info = lookup_number(num)
            if not info:
                ws.cell(row=row, column=7).value = "NOT_FOUND"
                continue

            ws.cell(row=row, column=2).value = info.get("name", "")
            ws.cell(row=row, column=3).value = info.get("father_name", "")
            ws.cell(row=row, column=4).value = info.get("address", "")
            ws.cell(row=row, column=5).value = info.get("email", "")
            ws.cell(row=row, column=6).value = info.get("alt_mobile", "")
            ws.cell(row=row, column=7).value = "FOUND"

        wb.save(temp_out)

        return FileResponse(
            temp_out,
            filename=f"Nextify_Output_{datetime.now().strftime("%H%M%S")}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)